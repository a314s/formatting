import openpyxl
import re
import tkinter as tk
from tkinter import filedialog
import os

def clean_text(original_text: str, options: dict) -> str:
    """
    Cleans text based on provided options dictionary.
    Keys in options dict correspond to checkboxes, e.g., options['remove_ellipsis'].
    """
    # 1. Strip leading/trailing whitespace
    text = original_text.strip()

    # 2. Conditionally remove all occurrences of '...'
    if options.get('remove_ellipsis', False):
        text = text.replace("...", "")

    # 3. Process quoted text (both single and double quotes).
    def process_quoted(match):
        quote_char = match.group(1)  # The quote symbol (single or double)
        content    = match.group(2)  # The text inside the quotes

        # Trim leading/trailing spaces inside the quotes
        content_stripped = content.strip()

        # If the content is "BOM" (any case), remove quotes entirely => BOM
        if content_stripped.upper() == "BOM":
            return "BOM"

        # Conditionally remove spaces from single letters in quotes
        if options.get('remove_spaces_quotes', False):
            tokens = content_stripped.split()
            # Check if content is not empty and all parts are single characters
            if content_stripped and all(len(t) == 1 for t in tokens):
                content_stripped = "".join(tokens)

        # Conditionally remove lone quotes (if content becomes empty after stripping/processing)
        if options.get('remove_lone_quotes', False) and not content_stripped:
             return "" # Return empty string if quote content is empty and option is set

        # Return with the original quote characters preserved, unless it was BOM or removed lone quote
        return f"{quote_char}{content_stripped}{quote_char}"

    print(f"DEBUG clean_text: Input='{original_text}', Options={options}") # DEBUG
    print(f"DEBUG clean_text: Before re.sub: text='{text}'") # DEBUG
    text = re.sub(r'(["\'])(.*?)(\1)', process_quoted, text)
    print(f"DEBUG clean_text: After re.sub: text='{text}'") # DEBUG

    # Handle unquoted single characters if the option is enabled
    if options.get('remove_spaces_unquoted', False):
        # Regex explanation:
        # (?<!["'])                     # Negative lookbehind: Ensure not preceded by a quote
        # (                             # Start capturing group 1
        #   (?:                         # Start non-capturing group for the sequence
        #     \b[a-zA-Z0-9]\b           # Match a single alphanumeric character (word boundary)
        #     (?:\s+                    # Match one or more spaces...
        #       (?![a-zA-Z0-9]\s*[a-zA-Z0-9]) # ...unless followed by another word (to avoid merging words)
        #       (?![.,;:!?])            # ...or followed by punctuation
        #     )
        #   )+                          # Repeat the single char + space pattern one or more times
        #   \b[a-zA-Z0-9]\b           # Match the final single alphanumeric character
        # )                             # End capturing group 1
        # (?![."'])                     # Negative lookahead: Ensure not followed by a quote or period
        #
        # This regex is complex and might need refinement based on edge cases.
        # It aims to find sequences of 2 or more single characters separated by spaces.
        # A simpler approach might be needed if this is too broad or misses cases.

        # Simpler approach: Find sequences of (single char + space) repeated 2+ times, followed by a single char.
        def remove_spaces_in_match(match):
            # Get the matched sequence (e.g., "P R 2 ")
            sequence = match.group(0)
            # Remove all spaces within the sequence
            return re.sub(r'\s+', '', sequence)

        # Find patterns like "X Y Z" or "A B C D" (at least 3 single chars)
        # \b([a-zA-Z0-9])\s+ means single char followed by space
        # (?: ... ){2,} means repeat that pattern at least twice
        # \b([a-zA-Z0-9])\b means end with a single char
        # We use a function replacement to avoid removing spaces between words accidentally
        text = re.sub(r'\b(?:[a-zA-Z0-9]\s+){2,}[a-zA-Z0-9]\b', remove_spaces_in_match, text)
        print(f"DEBUG clean_text: After remove_spaces_unquoted: text='{text}'") # DEBUG

    # Conditionally remove lone quotes that might remain after regex processing or were never paired
    if options.get('remove_lone_quotes', False):
         # Remove empty quotes first, then potentially single quotes if they remain
         text = text.replace('""', '').replace("''", '')
         # This part is tricky, might remove intended single quotes. Needs careful thought.
         # A more robust approach might involve checking quote balance.
         # For now, let's remove only explicitly empty pairs.

    # 4. Convert multiple spaces to single space (Generally always useful)
    text = re.sub(r'\s{2,}', ' ', text)

    # 5. Conditionally capitalize the first letter
    if options.get('capitalize_sentences', False) and text:
        text = text[0].upper() + text[1:]

    # 6. Conditionally ensure it ends with a period
    if options.get('add_periods', False) and text and not text.endswith("."):
        # Avoid adding period if the last char isn't suitable (e.g., another punctuation)
        # This is a basic check, more robust checks might be needed.
        if text[-1].isalnum() or text[-1] in ')]}"\'':
             text += "."

    return text

def remove_blank_rows(input_filename: str, sheet_name: str = None):
    """
    Removes all blank rows from the specified (or active) sheet in an Excel file.
    A row is considered blank if all cells in that row are empty.
    """
    # Load the workbook
    wb = openpyxl.load_workbook(input_filename)
    
    # If sheet_name is given, use that sheet; otherwise use active
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    
    # Find all blank rows (in reverse order to avoid index issues when deleting)
    blank_rows = []
    for row_idx in range(ws.max_row, 0, -1):
        is_blank = True
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None and str(cell.value).strip() != "":
                is_blank = False
                break
        if is_blank:
            blank_rows.append(row_idx)
    
    # Delete the blank rows
    for row_idx in blank_rows:
        ws.delete_rows(row_idx)
    
    # Save the workbook
    wb.save(input_filename)
    
    print(f"Removed {len(blank_rows)} blank rows from {input_filename}.")
    return len(blank_rows)

def process_excel(input_filename: str, options: dict, sheet_name: str = None):
    """
    Iterates down column A in the specified (or active) sheet.
    For each non-empty cell:
      - Clean the text using clean_text().
      - Write it back to the cell.
    Stops when two consecutive blank cells are found.
    """
    # Load the workbook
    wb = openpyxl.load_workbook(input_filename)
    
    # If sheet_name is given, use that sheet; otherwise use active
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    consecutive_blank_count = 0
    row = 1  # Start from the first row
    
    while True:
        cell = ws.cell(row=row, column=1)
        cell_value = cell.value
        
        # Check for blank (or None) cell
        if cell_value is None or str(cell_value).strip() == "":
            consecutive_blank_count += 1
            # If we've reached two consecutive blanks, stop
            if consecutive_blank_count == 2:
                break
        else:
            # Reset blank count
            consecutive_blank_count = 0

            # Convert cell_value to string and clean it using provided options
            text = str(cell_value)
            new_text = clean_text(text, options)

            # Update the cell if changed
            if new_text != text:
                cell.value = new_text
        
        row += 1  # Move to the next row
    
    # Option A: Overwrite the same file (make sure it's not open in Excel)
    wb.save(input_filename)

    # Option B: Save to a new file to avoid overwriting
    # new_filename = os.path.join(
    #     os.path.dirname(input_filename),
    #     "cleaned_" + os.path.basename(input_filename)
    # )
    # wb.save(new_filename)

    print(f"Finished processing {input_filename}.")

def main():
    """
    Main function that provides a GUI for selecting an Excel file and processing it.
    """
    # Create a hidden root window for the file dialog
    root = tk.Tk()
    root.withdraw()

    print("Please select an Excel file to process...")
    
    # Ask the user to select the Excel file
    input_filename = filedialog.askopenfilename(
        title="Select an Excel file",
        filetypes=[("Excel files", "*.xlsx *.xlsm *.xltx *.xltm")]
    )
    
    # If user cancels, just exit
    if not input_filename:
        print("No file selected. Exiting.")
        return
    
    # Ask user which operation to perform
    print("\nChoose an operation:")
    print("1. Clean text in cells")
    print("2. Remove blank rows")
    print("3. Clean text AND remove blank rows")
    choice = input("Enter your choice (1, 2, or 3): ")
    
    if choice == "1":
        # Process the chosen file - clean text
        process_excel(input_filename)
    elif choice == "2":
        # Remove blank rows from the chosen file
        removed = remove_blank_rows(input_filename)
        print(f"Removed {removed} blank rows from the file.")
    elif choice == "3":
        # Perform both operations
        print("Performing complete formatting...")
        process_excel(input_filename)  # First clean the text
        removed = remove_blank_rows(input_filename)  # Then remove blank rows
        print(f"Cleaned text and removed {removed} blank rows from the file.")
    else:
        print("Invalid choice. Exiting.")

if __name__ == "__main__":
    main()
