import openpyxl
from openpyxl import Workbook
import keyboard
import pyperclip
import time

# Create a new workbook and active sheet
title = input("Enter the title for the Excel sheet: ")
wb = Workbook()
ws = wb.active
ws.title = title

# Starting from the first row
row = 1

def add_text_to_excel():
    global row
    # Add a half-second delay to ensure the clipboard updates
    time.sleep(0.5)
    
    # Get the clipboard content
    selected_text = pyperclip.paste()
    
    # Add the clipboard content to the current row
    ws.cell(row=row, column=1).value = selected_text
    
    # Add a blank row after the text
    row += 1
    ws.cell(row=row, column=1).value = ''
    
    # Move to the next row
    row += 1

def save_and_exit():
    # Save the Excel file with the given title
    file_name = f"{title}.xlsx"
    wb.save(file_name)
    print(f"Excel file '{file_name}' saved.")
    exit()

# Set hotkeys
keyboard.add_hotkey('ctrl+c', add_text_to_excel)
keyboard.add_hotkey('esc', save_and_exit)

# Keep the script running
print("Script is running. Press Ctrl+C to add text from the clipboard, and Esc to save and exit.")
keyboard.wait('esc')
